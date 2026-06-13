#!/usr/bin/env python3
"""
Local optimizer for the Keltner Pine strategy used in this project.

What it does:
- Reads default/config values from a Pine file and an optional TradingView XLSX export.
- Pulls historical OHLC bars from Alpaca (using a user stored in local DB) or from a CSV file.
- Runs many backtests over random parameter combinations and ranks results.
- Saves a machine-readable report (JSON) and top combinations (CSV).

Notes:
- This is a pragmatic emulator for TradingView strategy behavior, not a bit-perfect clone.
- It models the broker emulator intrabar path assumption (open->high->low->close or open->low->high->close)
  and supports fixed SL/TP, forced SL/TP and trailing exits.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import os
import random
import re
import sys
from concurrent.futures import ProcessPoolExecutor
from copy import deepcopy
from dataclasses import asdict, dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import numpy as np
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))
DEFAULT_KELTNER_PINE = PROJECT_ROOT / "misc" / "strategies" / "keltner.pine"
DEFAULT_MACD_SMA_PINE = PROJECT_ROOT / "misc" / "strategies" / "MACD_SMA_strategy.pine"
DEFAULT_PINE = DEFAULT_KELTNER_PINE
DEFAULT_XLSX = PROJECT_ROOT / "misc" / "Keltner_channel_strategy_stocks_NYSE_TSM_2026-06-02.xlsx"
DEFAULT_REPORT = PROJECT_ROOT / "misc" / "optimizer_report.json"
DEFAULT_TOP_CSV = PROJECT_ROOT / "misc" / "optimizer_top.csv"
SUPPORTED_STRATEGIES = {"keltner", "macd_sma"}


@dataclass
class StrategyParams:
    trade_direction: str
    inner_kc_length: int
    inner_kc_mult: float
    outer_kc_length: int
    outer_kc_mult: float
    fixed_stop_loss_pct: float
    fixed_take_profit_pct: float
    forced_stop_loss_pct: float
    forced_take_profit_pct: float
    trailing_offset_ticks: int
    tick_size: float
    # When > 0 the trailing stop trails by this percent of price instead of a
    # fixed number of ticks. Keeps the trail proportional to the instrument so
    # winners on $100+ names are not cut at a few cents of give-back.
    trailing_offset_pct: float = 0.0
    macd_fast_length: int = 12
    macd_slow_length: int = 26
    macd_signal_length: int = 9
    macd_sma_length: int = 200
    max_intraday_loss_pct: float = 50.0


@dataclass
class BacktestConfig:
    initial_capital: float
    order_size_usd: float
    commission_pct: float
    timezone_name: str
    market_close_utc_hour: int = 20
    market_close_utc_minute: int = 30
    close_before_minutes: int = 4


@dataclass
class BacktestResult:
    params: StrategyParams
    final_equity: float
    net_profit: float
    return_pct: float
    total_trades: int
    winners: int
    losers: int
    win_rate_pct: float
    gross_profit: float
    gross_loss: float
    profit_factor: float
    sharpe: float
    max_drawdown: float
    max_drawdown_pct: float
    avg_bars_per_trade: float
    score: float
    trades: List[Dict[str, object]]


_WORKER_DF: Optional[pd.DataFrame] = None
_WORKER_CFG: Optional[BacktestConfig] = None
_WORKER_START: Optional[datetime] = None
_WORKER_END: Optional[datetime] = None
_WORKER_STRATEGY: str = "keltner"


def _safe_float(value: object, default: float) -> float:
    try:
        if value is None:
            return default
        return float(str(value).replace("%", "").replace(",", "").strip())
    except Exception:
        return default


def _safe_int(value: object, default: int) -> int:
    try:
        if value is None:
            return default
        return int(float(str(value).strip()))
    except Exception:
        return default


def parse_pine_defaults(pine_path: Path) -> Dict[str, object]:
    text = pine_path.read_text(encoding="utf-8")

    defaults: Dict[str, object] = {
        "trade_direction": "Both",
        "inner_kc_length": 13,
        "inner_kc_mult": 1.0,
        "outer_kc_length": 13,
        "outer_kc_mult": 2.0,
        "fixed_stop_loss_pct": 3.0,
        "fixed_take_profit_pct": 3.0,
        "forced_stop_loss_pct": 6.0,
        "forced_take_profit_pct": 6.0,
        "trailing_offset_ticks": 4,
        "trailing_offset_pct": 0.0,
        "initial_capital": 1000.0,
        "commission_pct": 0.04,
        "order_size_usd": 2000.0,
        "tick_size": 0.01,
        "macd_fast_length": 12,
        "macd_slow_length": 26,
        "macd_signal_length": 9,
        "macd_sma_length": 200,
        "max_intraday_loss_pct": 50.0,
    }

    # Parse strategy(...) line for initial capital and commission.
    strat_match = re.search(r"strategy\((.*?)\)", text, flags=re.S)
    if strat_match:
        s = strat_match.group(1)
        m_cap = re.search(r"initial_capital\s*=\s*([\d\.]+)", s)
        m_comm = re.search(r"commission_value\s*=\s*([\d\.]+)", s)
        if m_cap:
            defaults["initial_capital"] = _safe_float(m_cap.group(1), defaults["initial_capital"])
        if m_comm:
            defaults["commission_pct"] = _safe_float(m_comm.group(1), defaults["commission_pct"])

    def _extract_input_default(title: str, func: str = "input") -> Optional[str]:
        pat = rf"{func}\s*\(\s*([^,\)]+)\s*,.*?title\s*=\s*['\"]{re.escape(title)}['\"]"
        m = re.search(pat, text, flags=re.S)
        if not m:
            return None
        return m.group(1).strip().strip("'\"")

    mapping = [
        ("inner_kc_length", "Inner KC Length", "input", True),
        ("inner_kc_mult", "Inner KC Multiplier", "input", False),
        ("outer_kc_length", "Outer KC Length", "input", True),
        ("outer_kc_mult", "Outer KC Multiplier", "input", False),
        ("fixed_stop_loss_pct", "Fixed Stop Loss (%)", "input.float", False),
        ("fixed_take_profit_pct", "Fixed Take Profit (%)", "input.float", False),
        ("forced_stop_loss_pct", "Forced Stop Loss (%)", "input.float", False),
        ("forced_take_profit_pct", "Forced Take Profit (%)", "input.float", False),
    ]

    for key, title, func, is_int in mapping:
        v = _extract_input_default(title, func=func)
        if v is None:
            continue
        if is_int:
            defaults[key] = _safe_int(v, int(defaults[key]))
        else:
            defaults[key] = _safe_float(v, float(defaults[key]))

    m_trade = re.search(r"trade_direction\s*=\s*input\.string\((.*?)\)", text, flags=re.S)
    if m_trade:
        m_def = re.search(r"defval\s*=\s*\"([^\"]+)\"", m_trade.group(1))
        if m_def:
            defaults["trade_direction"] = m_def.group(1).strip()

    m_trail = re.search(r"trail_offset\s*=\s*([\d\.]+)", text)
    if m_trail:
        defaults["trailing_offset_ticks"] = _safe_int(m_trail.group(1), defaults["trailing_offset_ticks"])

    macd_mapping = [
        ("macd_fast_length", "MACD fast moving average", "input", True),
        ("macd_slow_length", "MACD slow moving average", "input", True),
        ("macd_signal_length", "MACD signal line moving average", "input", True),
        ("macd_sma_length", "Very slow moving average", "input", True),
        ("max_intraday_loss_pct", "Max Intraday Loss(%)", "input", False),
    ]
    for key, title, func, is_int in macd_mapping:
        v = _extract_input_default(title, func=func)
        if v is None:
            continue
        defaults[key] = _safe_int(v, int(defaults[key])) if is_int else _safe_float(v, float(defaults[key]))

    return defaults


def read_reference_xlsx(xlsx_path: Path) -> Dict[str, object]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    out: Dict[str, object] = {
        "properties": {},
        "performance": {},
        "trades_analysis": {},
        "risk_adjusted": {},
    }

    if "Properties" in wb.sheetnames:
        ws = wb["Properties"]
        for r in range(2, ws.max_row + 1):
            k = ws.cell(r, 1).value
            v = ws.cell(r, 2).value
            if k:
                out["properties"][str(k).strip()] = v

    def read_metric_sheet(sheet_name: str, target_key: str) -> None:
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]
        # first column metric name, second column "All USD", third "All %"
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if not name:
                continue
            usd_val = ws.cell(r, 2).value
            pct_val = ws.cell(r, 3).value
            out[target_key][str(name).strip()] = {"usd": usd_val, "pct": pct_val}

    read_metric_sheet("Performance", "performance")
    read_metric_sheet("Trades analysis", "trades_analysis")
    read_metric_sheet("Risk-adjusted performance", "risk_adjusted")
    return out


def parse_tv_datetime(dt_str: str, tz_name: str) -> datetime:
    # Example: "Jan 03, 2023, 16:30"
    naive = datetime.strptime(dt_str.strip(), "%b %d, %Y, %H:%M")
    # Keep dependency-light: use fixed offset from local timezone today as practical approximation.
    # For DST-accurate conversions, Python zoneinfo is used below when available.
    try:
        from zoneinfo import ZoneInfo

        return naive.replace(tzinfo=ZoneInfo(tz_name)).astimezone(timezone.utc)
    except Exception:
        # Fallback to local +02/+03 style through current system offset.
        local_off = datetime.now().astimezone().utcoffset() or timedelta(0)
        return naive.replace(tzinfo=timezone(local_off)).astimezone(timezone.utc)


def derive_reference_settings(
    pine_defaults: Dict[str, object],
    ref_props: Dict[str, object],
    timezone_name: str,
) -> Tuple[StrategyParams, BacktestConfig, str, str, datetime, datetime, str]:
    prop = ref_props

    # Symbol is usually like NYSE:TSM
    raw_symbol = str(prop.get("Symbol", "NYSE:TSM"))
    symbol = raw_symbol.split(":")[-1].strip().upper()

    timeframe = str(prop.get("Timeframe", "30 minutes")).strip()
    if timeframe.lower().startswith("30"):
        tf = "30Min"
    elif timeframe.lower().startswith("1 hour"):
        tf = "1Hour"
    else:
        tf = "30Min"

    backtest_range = str(prop.get("Backtesting range", "Jan 03, 2023, 16:30 — Jun 01, 2026, 22:30"))
    if "—" in backtest_range:
        start_s, end_s = [x.strip() for x in backtest_range.split("—", 1)]
    elif "-" in backtest_range:
        start_s, end_s = [x.strip() for x in backtest_range.split("-", 1)]
    else:
        start_s, end_s = "Jan 03, 2023, 16:30", "Jun 01, 2026, 22:30"

    start_utc = parse_tv_datetime(start_s, timezone_name)
    end_utc = parse_tv_datetime(end_s, timezone_name)

    params = StrategyParams(
        trade_direction=str(prop.get("Trade Direction", pine_defaults["trade_direction"])),
        inner_kc_length=_safe_int(prop.get("Inner KC Length"), int(pine_defaults["inner_kc_length"])),
        inner_kc_mult=_safe_float(prop.get("Inner KC Multiplier"), float(pine_defaults["inner_kc_mult"])),
        outer_kc_length=_safe_int(prop.get("Outer KC Length"), int(pine_defaults["outer_kc_length"])),
        outer_kc_mult=_safe_float(prop.get("Outer KC Multiplier"), float(pine_defaults["outer_kc_mult"])),
        fixed_stop_loss_pct=_safe_float(prop.get("Fixed Stop Loss (%)"), float(pine_defaults["fixed_stop_loss_pct"])),
        fixed_take_profit_pct=_safe_float(prop.get("Fixed Take Profit (%)"), float(pine_defaults["fixed_take_profit_pct"])),
        forced_stop_loss_pct=_safe_float(prop.get("Forced Stop Loss (%)"), float(pine_defaults["forced_stop_loss_pct"])),
        forced_take_profit_pct=_safe_float(prop.get("Forced Take Profit (%)"), float(pine_defaults["forced_take_profit_pct"])),
        trailing_offset_ticks=int(pine_defaults["trailing_offset_ticks"]),
        tick_size=_safe_float(prop.get("Tick size"), float(pine_defaults["tick_size"])),
        trailing_offset_pct=float(pine_defaults.get("trailing_offset_pct", 0.0)),
    )

    cfg = BacktestConfig(
        initial_capital=_safe_float(prop.get("Initial capital"), float(pine_defaults["initial_capital"])),
        order_size_usd=_safe_float(prop.get("Order size"), float(pine_defaults["order_size_usd"])),
        commission_pct=_safe_float(prop.get("Commission"), float(pine_defaults["commission_pct"])),
        timezone_name=timezone_name,
    )

    return params, cfg, symbol, tf, start_utc, end_utc, raw_symbol


def fetch_bars_alpaca(
    symbol: str,
    timeframe: str,
    start_utc: datetime,
    end_utc: datetime,
    username: Optional[str],
    feed: str,
) -> pd.DataFrame:
    load_dotenv(PROJECT_ROOT / ".env")

    from flask import Flask

    from alpaca_api import LegacyCompatibleAlpacaClient
    from models import User, db
    from utils import decrypt_data

    app = Flask(__name__)
    app.config["SQLALCHEMY_DATABASE_URI"] = __import__("os").getenv("DATABASE_URL", "sqlite:///app.db")
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
    db.init_app(app)

    with app.app_context():
        if username:
            user = User.query.filter_by(username=username).first()
        else:
            user = User.query.filter_by(is_superuser=True).first() or User.query.first()
        if not user:
            raise RuntimeError("No user found in local DB to decrypt Alpaca credentials.")

        key = decrypt_data(user.encrypted_alpaca_key)
        secret = decrypt_data(user.encrypted_alpaca_secret)
        if not key or not secret:
            raise RuntimeError(f"User '{user.username}' has no usable Alpaca credentials.")

        base_url = __import__("os").getenv("ALPACA_API_BASE_URL", "https://paper-api.alpaca.markets")
        api = LegacyCompatibleAlpacaClient(key, secret, base_url)

        bars = api.get_bars(
            symbol,
            timeframe,
            start=start_utc.isoformat().replace("+00:00", "Z"),
            end=end_utc.isoformat().replace("+00:00", "Z"),
            adjustment="raw",
            feed=feed,
        ).df

    if bars.empty:
        raise RuntimeError("No bars returned from Alpaca. Check symbol/timeframe/feed/range.")

    # Normalize columns and ordering.
    bars = bars.copy()
    bars = bars[["open", "high", "low", "close", "volume"]]
    bars = bars.sort_index()
    bars = bars[~bars.index.duplicated(keep="first")]
    bars.index = pd.to_datetime(bars.index, utc=True)
    return bars


def load_bars_csv(csv_path: Path) -> pd.DataFrame:
    df = pd.read_csv(csv_path)
    needed = {"timestamp", "open", "high", "low", "close"}
    if not needed.issubset(set(df.columns)):
        raise ValueError(f"CSV must include columns: {sorted(needed)}")
    df["timestamp"] = pd.to_datetime(df["timestamp"], utc=True)
    df = df.set_index("timestamp").sort_index()
    if "volume" not in df.columns:
        df["volume"] = 0
    return df[["open", "high", "low", "close", "volume"]]


def filter_session(df: pd.DataFrame, mode: str) -> pd.DataFrame:
    mode = (mode or "regular").lower()
    if mode == "all":
        return df
    if len(df.index) > 1:
        min_delta = df.index.to_series().diff().dropna().min()
        if pd.notna(min_delta) and min_delta >= pd.Timedelta(hours=23):
            return df if mode == "regular" else df.iloc[0:0]
    ny = df.tz_convert("America/New_York")
    minutes = ny.index.hour * 60 + ny.index.minute
    regular_mask = (minutes >= (9 * 60 + 30)) & (minutes < (16 * 60))
    if mode == "regular":
        return df[regular_mask]
    if mode == "extended":
        return df[~regular_mask]
    raise ValueError(f"Unknown session mode: {mode}")

def normalize_timeframe_token(value: str) -> str:
    token = str(value or "").strip().lower().replace(" ", "")
    m = re.fullmatch(r"(\d+)(min|m|minute|minutes|hour|h|hours|day|d|days|week|w|weeks)", token)
    if not m:
        raise ValueError(f"Unsupported timeframe: {value}")
    amount = int(m.group(1))
    unit = m.group(2)
    if amount < 1:
        raise ValueError("Timeframe amount must be >= 1")
    if unit in ("min", "m", "minute", "minutes"):
        return f"{amount}Min"
    if unit in ("hour", "h", "hours"):
        return f"{amount}Hour"
    if unit in ("day", "d", "days"):
        return f"{amount}Day"
    if unit in ("week", "w", "weeks"):
        return f"{amount}Week"
    raise ValueError(f"Unsupported timeframe: {value}")

def timeframe_seconds(label: str) -> int:
    normalized = normalize_timeframe_token(label)
    m = re.fullmatch(r"(\d+)(Min|Hour|Day|Week)", normalized)
    amount = int(m.group(1))
    unit = m.group(2)
    multipliers = {"Min": 60, "Hour": 3600, "Day": 86400, "Week": 604800}
    return amount * multipliers[unit]

def parse_timeframes_arg(value: Optional[str], fallback: str) -> List[str]:
    raw = value or fallback
    labels = [normalize_timeframe_token(part) for part in str(raw).split(",") if part.strip()]
    if not labels:
        labels = [normalize_timeframe_token(fallback)]
    out = []
    seen = set()
    for label in labels:
        if label not in seen:
            out.append(label)
            seen.add(label)
    out.sort(key=timeframe_seconds)
    return out

def choose_fetch_timeframe(labels: List[str]) -> str:
    minute_amounts = []
    has_hour = False
    has_day_or_week = False
    for label in labels:
        m = re.fullmatch(r"(\d+)(Min|Hour|Day|Week)", normalize_timeframe_token(label))
        amount = int(m.group(1))
        unit = m.group(2)
        if unit == "Min":
            minute_amounts.append(amount)
        elif unit == "Hour":
            has_hour = True
        else:
            has_day_or_week = True
    if minute_amounts:
        from math import gcd
        base = minute_amounts[0]
        for amount in minute_amounts[1:]:
            base = gcd(base, amount)
        return f"{max(1, base)}Min"
    if has_hour:
        return "1Hour"
    if has_day_or_week:
        return "1Day"
    return labels[0]

def timeframe_to_pandas_rule(label: str) -> str:
    normalized = normalize_timeframe_token(label)
    m = re.fullmatch(r"(\d+)(Min|Hour|Day|Week)", normalized)
    amount = int(m.group(1))
    unit = m.group(2)
    suffix = {"Min": "min", "Hour": "h", "Day": "D", "Week": "W"}[unit]
    return f"{amount}{suffix}"

def resample_bars(df: pd.DataFrame, timeframe: str) -> pd.DataFrame:
    rule = timeframe_to_pandas_rule(timeframe)
    agg = {
        "open": "first",
        "high": "max",
        "low": "min",
        "close": "last",
        "volume": "sum",
    }
    out = df.sort_index().resample(rule, label="right", closed="right").agg(agg)
    return out.dropna(subset=["open", "high", "low", "close"])


def ema(series: pd.Series, length: int) -> pd.Series:
    return series.ewm(span=length, adjust=False, min_periods=length).mean()


def sma(series: pd.Series, length: int) -> pd.Series:
    return series.rolling(window=length, min_periods=length).mean()


def crossed_above(series: pd.Series, level: float = 0.0) -> pd.Series:
    return (series.shift(1) <= level) & (series > level)


def crossed_below(series: pd.Series, level: float = 0.0) -> pd.Series:
    return (series.shift(1) >= level) & (series < level)


def true_range(df: pd.DataFrame) -> pd.Series:
    prev_close = df["close"].shift(1)
    h_l = df["high"] - df["low"]
    h_pc = (df["high"] - prev_close).abs()
    l_pc = (df["low"] - prev_close).abs()
    return pd.concat([h_l, h_pc, l_pc], axis=1).max(axis=1)


def keltner_channel(df: pd.DataFrame, length: int, mult: float) -> Tuple[pd.Series, pd.Series, pd.Series]:
    basis = ema(df["close"], length)
    span = ema(true_range(df), length)
    upper = basis + span * mult
    lower = basis - span * mult
    return basis, upper, lower


def infer_path(open_p: float, high_p: float, low_p: float, close_p: float) -> List[float]:
    # TradingView historical emulator heuristic.
    if abs(high_p - open_p) <= abs(open_p - low_p):
        return [open_p, high_p, low_p, close_p]
    return [open_p, low_p, high_p, close_p]


def crosses_up(a: float, b: float, level: float) -> bool:
    return min(a, b) <= level <= max(a, b) and b >= a and b >= level


def crosses_down(a: float, b: float, level: float) -> bool:
    return min(a, b) <= level <= max(a, b) and b <= a and b <= level


def long_intrabar_exit(
    path: List[float],
    activation: float,
    forced_stop: float,
    forced_tp: float,
    trail_offset: float,
    trail_active: bool,
    trail_stop: Optional[float],
    allow_trail: bool,
) -> Tuple[Optional[float], Optional[str], bool, Optional[float]]:
    # Returns: (exit_price, reason, new_trail_active, new_trail_stop)
    active = trail_active
    stop = trail_stop

    for i in range(len(path) - 1):
        a, b = path[i], path[i + 1]

        # Gap/segment hit checks for forced orders first (OCA stop/limit style).
        if b >= a:
            # Up segment: forced TP can hit.
            if a <= forced_tp <= b:
                return forced_tp, "Forced TP Long", active, stop
        else:
            # Down segment: forced SL can hit.
            if b <= forced_stop <= a:
                return forced_stop, "Forced SL Long", active, stop

        if not allow_trail:
            continue

        # Trailing activation/update.
        if not active:
            # Activation when price crosses upward the activation level.
            if b >= a and a <= activation <= b:
                active = True
                stop = activation - trail_offset
            elif a >= activation:
                active = True
                stop = a - trail_offset

        if active:
            if b >= a:
                # Favorable move: update stop with new highs reached in segment.
                candidate = b - trail_offset
                stop = candidate if stop is None else max(stop, candidate)
            else:
                # Adverse move: stop may trigger.
                if stop is not None and b <= stop <= a:
                    return stop, "Trailing Exit Long", active, stop

    return None, None, active, stop


def short_intrabar_exit(
    path: List[float],
    activation: float,
    forced_stop: float,
    forced_tp: float,
    trail_offset: float,
    trail_active: bool,
    trail_stop: Optional[float],
    allow_trail: bool,
) -> Tuple[Optional[float], Optional[str], bool, Optional[float]]:
    active = trail_active
    stop = trail_stop

    for i in range(len(path) - 1):
        a, b = path[i], path[i + 1]

        if b <= a:
            # Down segment: forced TP can hit.
            if b <= forced_tp <= a:
                return forced_tp, "Forced TP Short", active, stop
        else:
            # Up segment: forced SL can hit.
            if a <= forced_stop <= b:
                return forced_stop, "Forced SL Short", active, stop

        if not allow_trail:
            continue

        if not active:
            # Activation when price crosses downward activation.
            if b <= a and b <= activation <= a:
                active = True
                stop = activation + trail_offset
            elif a <= activation:
                active = True
                stop = a + trail_offset

        if active:
            if b <= a:
                candidate = b + trail_offset
                stop = candidate if stop is None else min(stop, candidate)
            else:
                if stop is not None and a <= stop <= b:
                    return stop, "Trailing Exit Short", active, stop

    return None, None, active, stop


def in_preclose_window(ts_utc: pd.Timestamp, cfg: BacktestConfig) -> bool:
    # Approximation for the Pine time-window logic.
    mc = ts_utc.replace(hour=cfg.market_close_utc_hour, minute=cfg.market_close_utc_minute, second=0, microsecond=0)
    start = mc - timedelta(minutes=cfg.close_before_minutes)
    return start <= ts_utc < mc


# ---------------------------------------------------------------------------
# Optional Numba JIT fast path for the Keltner simulation.
#
# The bar loop is sequential and branch-heavy, so it dominates optimizer runtime.
# The functions below are written in a Numba-compatible style (numpy + scalars,
# no pandas/objects). When numba is importable they are JIT-compiled for a large
# speedup; otherwise they run as plain Python with identical semantics. The
# original backtest() remains the reference implementation and the fast path is
# only used after parity is verified (see tests/test_pine_optimizer_numba.py).
# Set STRATEGY_DISABLE_NUMBA=1 to force the reference path.
# ---------------------------------------------------------------------------
try:
    if os.getenv("STRATEGY_DISABLE_NUMBA", "").strip().lower() in ("1", "true", "yes", "y"):
        raise ImportError("numba disabled via STRATEGY_DISABLE_NUMBA")
    from numba import njit as _njit  # type: ignore

    NUMBA_AVAILABLE = True
except Exception:  # pragma: no cover - environment without numba
    NUMBA_AVAILABLE = False

    def _njit(*args, **kwargs):  # type: ignore
        def _decorator(func):
            return func
        if args and callable(args[0]):
            return args[0]
        return _decorator


@_njit(cache=True)
def _long_intrabar_exit_core(p0, p1, p2, p3, activation, forced_stop, forced_tp,
                             trail_offset, trail_active, has_stop, stop, allow_trail):
    # Returns: (hit, exit_price, reason_code, active, has_stop, stop)
    active = trail_active
    hs = has_stop
    st = stop
    path = (p0, p1, p2, p3)
    for i in range(3):
        a = path[i]
        b = path[i + 1]
        if b >= a:
            if a <= forced_tp <= b:
                return 1, forced_tp, 5, active, hs, st  # Forced TP Long
        else:
            if b <= forced_stop <= a:
                return 1, forced_stop, 6, active, hs, st  # Forced SL Long
        if allow_trail == 0:
            continue
        if active == 0:
            if b >= a and a <= activation <= b:
                active = 1
                st = activation - trail_offset
                hs = 1
            elif a >= activation:
                active = 1
                st = a - trail_offset
                hs = 1
        if active == 1:
            if b >= a:
                candidate = b - trail_offset
                if hs == 0:
                    st = candidate
                    hs = 1
                elif candidate > st:
                    st = candidate
            else:
                if hs == 1 and b <= st <= a:
                    return 1, st, 9, active, hs, st  # Trailing Exit Long
    return 0, 0.0, 0, active, hs, st


@_njit(cache=True)
def _short_intrabar_exit_core(p0, p1, p2, p3, activation, forced_stop, forced_tp,
                              trail_offset, trail_active, has_stop, stop, allow_trail):
    active = trail_active
    hs = has_stop
    st = stop
    path = (p0, p1, p2, p3)
    for i in range(3):
        a = path[i]
        b = path[i + 1]
        if b <= a:
            if b <= forced_tp <= a:
                return 1, forced_tp, 7, active, hs, st  # Forced TP Short
        else:
            if a <= forced_stop <= b:
                return 1, forced_stop, 8, active, hs, st  # Forced SL Short
        if allow_trail == 0:
            continue
        if active == 0:
            if b <= a and b <= activation <= a:
                active = 1
                st = activation + trail_offset
                hs = 1
            elif a <= activation:
                active = 1
                st = a + trail_offset
                hs = 1
        if active == 1:
            if b <= a:
                candidate = b + trail_offset
                if hs == 0:
                    st = candidate
                    hs = 1
                elif candidate < st:
                    st = candidate
            else:
                if hs == 1 and a <= st <= b:
                    return 1, st, 10, active, hs, st  # Trailing Exit Short
    return 0, 0.0, 0, active, hs, st


@_njit(cache=True)
def _simulate_keltner_core(o, h, l, c, mid, uin, lin, preclose,
                           long_allowed, short_allowed, order_size_usd, commission_pct,
                           fixed_sl_pct, fixed_tp_pct, forced_sl_pct, forced_tp_pct,
                           trailing_offset_ticks, tick_size, trailing_offset_pct,
                           initial_capital):
    n = o.shape[0]
    ent_idx = np.empty(n, dtype=np.int64)
    ex_idx = np.empty(n, dtype=np.int64)
    ent_px = np.empty(n, dtype=np.float64)
    ex_px = np.empty(n, dtype=np.float64)
    qty_arr = np.empty(n, dtype=np.int64)
    side_arr = np.empty(n, dtype=np.int64)
    reason_arr = np.empty(n, dtype=np.int64)
    pnl_arr = np.empty(n, dtype=np.float64)
    pnlpct_arr = np.empty(n, dtype=np.float64)
    equity_arr = np.empty(n + 1, dtype=np.float64)
    equity_arr[0] = initial_capital
    eq_count = 1
    tcount = 0

    equity = initial_capital
    position = 0
    entry_price = 0.0
    qty = 0
    entry_index = -1
    trail_active = 0
    has_stop = 0
    trail_stop = 0.0

    for i in range(1, n):
        entered_this_bar = False

        if position == 0:
            long_cond = (c[i - 1] <= lin[i - 1]) and (c[i] > lin[i]) and (c[i] < mid[i])
            short_cond = (c[i - 1] >= uin[i - 1]) and (c[i] < uin[i]) and (c[i] > mid[i])
            entry_side = 0
            if long_allowed == 1 and long_cond:
                entry_side = 1
            elif short_allowed == 1 and short_cond:
                entry_side = -1
            if entry_side != 0:
                q = int(math.floor(order_size_usd / c[i]))
                if q > 0:
                    position = entry_side
                    entry_price = c[i]
                    qty = q
                    entry_index = i
                    trail_active = 0
                    has_stop = 0
                    trail_stop = 0.0
                    entered_this_bar = True

        if position != 0 and not entered_this_bar:
            exit_price = 0.0
            has_exit = 0
            reason_code = 0

            if trailing_offset_pct > 0:
                trail_offset = c[i] * trailing_offset_pct / 100.0
            else:
                trail_offset = trailing_offset_ticks * tick_size

            oi = o[i]
            hi = h[i]
            li = l[i]
            ci = c[i]
            if abs(hi - oi) <= abs(oi - li):
                p0, p1, p2, p3 = oi, hi, li, ci
            else:
                p0, p1, p2, p3 = oi, li, hi, ci

            fixed_code = 0
            if position > 0:
                fixed_stop = entry_price * (1 - fixed_sl_pct / 100.0)
                fixed_tp = entry_price * (1 + fixed_tp_pct / 100.0)
                forced_stop = entry_price * (1 - forced_sl_pct / 100.0)
                forced_tp = entry_price * (1 + forced_tp_pct / 100.0)
                if li <= fixed_stop:
                    fixed_code = 1
                if hi >= fixed_tp:
                    fixed_code = 2
                allow_trail = 1 if fixed_code == 0 else 0
                hit, ex_p, rc, trail_active, has_stop, trail_stop = _long_intrabar_exit_core(
                    p0, p1, p2, p3, mid[i], forced_stop, forced_tp, trail_offset,
                    trail_active, has_stop, trail_stop, allow_trail,
                )
                if hit == 1:
                    has_exit = 1
                    exit_price = ex_p
                    reason_code = rc
                if has_exit == 0 and preclose[i] == 1 and ci > entry_price:
                    has_exit = 1
                    exit_price = ci
                    reason_code = 5  # Forced TP Long
            else:
                fixed_stop = entry_price * (1 + fixed_sl_pct / 100.0)
                fixed_tp = entry_price * (1 - fixed_tp_pct / 100.0)
                forced_stop = entry_price * (1 + forced_sl_pct / 100.0)
                forced_tp = entry_price * (1 - forced_tp_pct / 100.0)
                if hi >= fixed_stop:
                    fixed_code = 3
                if li <= fixed_tp:
                    fixed_code = 4
                allow_trail = 1 if fixed_code == 0 else 0
                hit, ex_p, rc, trail_active, has_stop, trail_stop = _short_intrabar_exit_core(
                    p0, p1, p2, p3, mid[i], forced_stop, forced_tp, trail_offset,
                    trail_active, has_stop, trail_stop, allow_trail,
                )
                if hit == 1:
                    has_exit = 1
                    exit_price = ex_p
                    reason_code = rc
                if has_exit == 0 and preclose[i] == 1 and ci < entry_price:
                    has_exit = 1
                    exit_price = ci
                    reason_code = 7  # Forced TP Short

            if has_exit == 0 and fixed_code != 0:
                has_exit = 1
                exit_price = ci
                reason_code = fixed_code

            if has_exit == 1:
                notional_entry = qty * entry_price
                notional_exit = qty * exit_price
                fees = (notional_entry + notional_exit) * (commission_pct / 100.0)
                pnl = (exit_price - entry_price) * qty * position - fees
                pnl_pct = (pnl / notional_entry) * 100.0 if notional_entry != 0 else 0.0

                equity += pnl
                equity_arr[eq_count] = equity
                eq_count += 1

                ent_idx[tcount] = entry_index
                ex_idx[tcount] = i
                ent_px[tcount] = entry_price
                ex_px[tcount] = exit_price
                qty_arr[tcount] = qty
                side_arr[tcount] = position
                reason_arr[tcount] = reason_code
                pnl_arr[tcount] = pnl
                pnlpct_arr[tcount] = pnl_pct
                tcount += 1

                position = 0
                entry_price = 0.0
                qty = 0
                entry_index = -1
                trail_active = 0
                has_stop = 0
                trail_stop = 0.0

    if position != 0 and qty > 0:
        final_price = c[n - 1]
        notional_entry = qty * entry_price
        notional_exit = qty * final_price
        fees = (notional_entry + notional_exit) * (commission_pct / 100.0)
        pnl = (final_price - entry_price) * qty * position - fees
        pnl_pct = (pnl / notional_entry) * 100.0 if notional_entry != 0 else 0.0
        equity += pnl
        equity_arr[eq_count] = equity
        eq_count += 1
        ent_idx[tcount] = entry_index
        ex_idx[tcount] = n - 1
        ent_px[tcount] = entry_price
        ex_px[tcount] = final_price
        qty_arr[tcount] = qty
        side_arr[tcount] = position
        reason_arr[tcount] = 11  # Final Close
        pnl_arr[tcount] = pnl
        pnlpct_arr[tcount] = pnl_pct
        tcount += 1

    return (tcount, ent_idx, ex_idx, ent_px, ex_px, qty_arr, side_arr, reason_arr,
            pnl_arr, pnlpct_arr, eq_count, equity_arr)


_REASON_CODE_TO_TEXT = {
    1: "Fixed Stop Loss (Long)", 2: "Fixed Take Profit (Long)",
    3: "Fixed Stop Loss (Short)", 4: "Fixed Take Profit (Short)",
    5: "Forced TP Long", 6: "Forced SL Long",
    7: "Forced TP Short", 8: "Forced SL Short",
    9: "Trailing Exit Long", 10: "Trailing Exit Short",
    11: "Final Close",
}


def _preclose_flags(idx, cfg: BacktestConfig) -> np.ndarray:
    # Vectorized equivalent of in_preclose_window over the whole index. The window
    # is [market_close - close_before, market_close) on the same day, which for
    # bar-aligned timestamps reduces to a minute-of-day comparison.
    minutes_of_day = np.asarray(idx.hour) * 60 + np.asarray(idx.minute)
    close_min = cfg.market_close_utc_hour * 60 + cfg.market_close_utc_minute
    start_min = close_min - cfg.close_before_minutes
    return ((minutes_of_day >= start_min) & (minutes_of_day < close_min)).astype(np.int64)


def backtest_fast(df: pd.DataFrame, params: StrategyParams, cfg: BacktestConfig, start_utc: datetime, end_utc: datetime) -> BacktestResult:
    df = df[(df.index >= pd.Timestamp(start_utc)) & (df.index <= pd.Timestamp(end_utc))].copy()
    if len(df) < 100:
        raise RuntimeError("Not enough bars for backtest after date filtering.")

    mid_inner, up_inner, low_inner = keltner_channel(df, params.inner_kc_length, params.inner_kc_mult)
    df["mid_inner"] = mid_inner
    df["up_inner"] = up_inner
    df["low_inner"] = low_inner
    df = df.dropna()
    if df.empty:
        raise RuntimeError("Indicator warm-up removed all bars; widen date range.")

    idx = df.index
    o = df["open"].to_numpy(dtype=float)
    h = df["high"].to_numpy(dtype=float)
    l = df["low"].to_numpy(dtype=float)
    c = df["close"].to_numpy(dtype=float)
    mid = df["mid_inner"].to_numpy(dtype=float)
    uin = df["up_inner"].to_numpy(dtype=float)
    lin = df["low_inner"].to_numpy(dtype=float)
    preclose = _preclose_flags(idx, cfg)

    long_allowed = 1 if params.trade_direction in ("Both", "Long Only") else 0
    short_allowed = 1 if params.trade_direction in ("Both", "Short Only") else 0

    (tcount, ent_idx, ex_idx, ent_px, ex_px, qty_arr, side_arr, reason_arr,
     pnl_arr, pnlpct_arr, eq_count, equity_arr) = _simulate_keltner_core(
        o, h, l, c, mid, uin, lin, preclose,
        long_allowed, short_allowed, float(cfg.order_size_usd), float(cfg.commission_pct),
        float(params.fixed_stop_loss_pct), float(params.fixed_take_profit_pct),
        float(params.forced_stop_loss_pct), float(params.forced_take_profit_pct),
        int(params.trailing_offset_ticks), float(params.tick_size), float(params.trailing_offset_pct),
        float(cfg.initial_capital),
    )

    winners = 0
    losers = 0
    gross_profit = 0.0
    gross_loss = 0.0
    trades: List[Dict[str, object]] = []
    trade_returns: List[float] = []
    trade_bars: List[int] = []
    for t in range(tcount):
        pnl = float(pnl_arr[t])
        if pnl >= 0:
            gross_profit += pnl
            winners += 1
        else:
            gross_loss += abs(pnl)
            losers += 1
        ei = int(ent_idx[t])
        xi = int(ex_idx[t])
        bars_held = xi - ei
        trade_returns.append(float(pnlpct_arr[t]))
        trade_bars.append(bars_held)
        trades.append({
            "entry_time": idx[ei].isoformat() if ei >= 0 else None,
            "exit_time": idx[xi].isoformat(),
            "side": "long" if side_arr[t] > 0 else "short",
            "entry_price": round(float(ent_px[t]), 6),
            "exit_price": round(float(ex_px[t]), 6),
            "qty": int(qty_arr[t]),
            "pnl": round(pnl, 2),
            "pnl_pct": round(float(pnlpct_arr[t]), 4),
            "reason": _REASON_CODE_TO_TEXT.get(int(reason_arr[t]), "Unknown"),
            "bars_held": bars_held,
        })

    equity_marks = [float(x) for x in equity_arr[:eq_count]]
    equity = equity_marks[-1] if equity_marks else cfg.initial_capital
    return _finalize_result(
        params, cfg, equity, equity_marks, winners, losers,
        gross_profit, gross_loss, trade_returns, trade_bars, trades,
    )


def backtest(df: pd.DataFrame, params: StrategyParams, cfg: BacktestConfig, start_utc: datetime, end_utc: datetime) -> BacktestResult:
    df = df[(df.index >= pd.Timestamp(start_utc)) & (df.index <= pd.Timestamp(end_utc))].copy()
    if len(df) < 100:
        raise RuntimeError("Not enough bars for backtest after date filtering.")

    mid_inner, up_inner, low_inner = keltner_channel(df, params.inner_kc_length, params.inner_kc_mult)
    mid_outer, up_outer, low_outer = keltner_channel(df, params.outer_kc_length, params.outer_kc_mult)

    df["mid_inner"] = mid_inner
    df["up_inner"] = up_inner
    df["low_inner"] = low_inner
    df["mid_outer"] = mid_outer
    df["up_outer"] = up_outer
    df["low_outer"] = low_outer
    df = df.dropna()

    if df.empty:
        raise RuntimeError("Indicator warm-up removed all bars; widen date range.")

    idx = df.index
    o = df["open"].to_numpy(dtype=float)
    h = df["high"].to_numpy(dtype=float)
    l = df["low"].to_numpy(dtype=float)
    c = df["close"].to_numpy(dtype=float)
    mid = df["mid_inner"].to_numpy(dtype=float)
    uin = df["up_inner"].to_numpy(dtype=float)
    lin = df["low_inner"].to_numpy(dtype=float)

    equity = cfg.initial_capital
    equity_marks: List[float] = [equity]

    position = 0  # +1 long, -1 short, 0 flat
    entry_price = 0.0
    qty = 0
    entry_index = -1
    trail_active = False
    trail_stop: Optional[float] = None

    gross_profit = 0.0
    gross_loss = 0.0
    commission_paid = 0.0
    winners = 0
    losers = 0

    trade_returns: List[float] = []
    trade_bars: List[int] = []
    trades: List[Dict[str, object]] = []
    entry_time: Optional[pd.Timestamp] = None

    for i in range(1, len(df)):
        ts = idx[i]
        entered_this_bar = False

        # Entry logic first, because Pine evaluates entries before exit blocks.
        if position == 0:
            long_allowed = params.trade_direction in ("Both", "Long Only")
            short_allowed = params.trade_direction in ("Both", "Short Only")

            long_cond = c[i - 1] <= lin[i - 1] and c[i] > lin[i] and c[i] < mid[i]
            short_cond = c[i - 1] >= uin[i - 1] and c[i] < uin[i] and c[i] > mid[i]

            entry_side = 0
            if long_allowed and long_cond:
                entry_side = 1
            elif short_allowed and short_cond:
                entry_side = -1

            if entry_side != 0:
                q = math.floor(cfg.order_size_usd / c[i])
                if q > 0:
                    position = entry_side
                    entry_price = c[i]
                    qty = q
                    entry_index = i
                    entry_time = idx[i]
                    trail_active = False
                    trail_stop = None
                    entered_this_bar = True

        # Exit logic for positions that were already open before this bar.
        if position != 0 and not entered_this_bar:
            exit_price: Optional[float] = None
            exit_reason: Optional[str] = None

            fixed_reason: Optional[str] = None
            path = infer_path(o[i], h[i], l[i], c[i])
            if params.trailing_offset_pct > 0:
                trail_offset = c[i] * params.trailing_offset_pct / 100.0
            else:
                trail_offset = params.trailing_offset_ticks * params.tick_size

            if position > 0:
                fixed_stop = entry_price * (1 - params.fixed_stop_loss_pct / 100.0)
                fixed_tp = entry_price * (1 + params.fixed_take_profit_pct / 100.0)
                forced_stop = entry_price * (1 - params.forced_stop_loss_pct / 100.0)
                forced_tp = entry_price * (1 + params.forced_take_profit_pct / 100.0)

                if l[i] <= fixed_stop:
                    fixed_reason = "Fixed Stop Loss (Long)"
                if h[i] >= fixed_tp:
                    fixed_reason = "Fixed Take Profit (Long)"

                ex_p, ex_r, trail_active, trail_stop = long_intrabar_exit(
                    path=path,
                    activation=mid[i],
                    forced_stop=forced_stop,
                    forced_tp=forced_tp,
                    trail_offset=trail_offset,
                    trail_active=trail_active,
                    trail_stop=trail_stop,
                    allow_trail=(fixed_reason is None),
                )
                if ex_p is not None:
                    exit_price, exit_reason = ex_p, ex_r

                if exit_price is None and in_preclose_window(ts, cfg) and c[i] > entry_price:
                    exit_price = c[i]
                    exit_reason = "Forced TP Long"

            else:
                fixed_stop = entry_price * (1 + params.fixed_stop_loss_pct / 100.0)
                fixed_tp = entry_price * (1 - params.fixed_take_profit_pct / 100.0)
                forced_stop = entry_price * (1 + params.forced_stop_loss_pct / 100.0)
                forced_tp = entry_price * (1 - params.forced_take_profit_pct / 100.0)

                if h[i] >= fixed_stop:
                    fixed_reason = "Fixed Stop Loss (Short)"
                if l[i] <= fixed_tp:
                    fixed_reason = "Fixed Take Profit (Short)"

                ex_p, ex_r, trail_active, trail_stop = short_intrabar_exit(
                    path=path,
                    activation=mid[i],
                    forced_stop=forced_stop,
                    forced_tp=forced_tp,
                    trail_offset=trail_offset,
                    trail_active=trail_active,
                    trail_stop=trail_stop,
                    allow_trail=(fixed_reason is None),
                )
                if ex_p is not None:
                    exit_price, exit_reason = ex_p, ex_r

                if exit_price is None and in_preclose_window(ts, cfg) and c[i] < entry_price:
                    exit_price = c[i]
                    exit_reason = "Forced TP Short"

            # strategy.close based fixed exits happen at bar close.
            if exit_price is None and fixed_reason is not None:
                exit_price = c[i]
                exit_reason = fixed_reason

            if exit_price is not None:
                notional_entry = qty * entry_price
                notional_exit = qty * exit_price
                fees = (notional_entry + notional_exit) * (cfg.commission_pct / 100.0)
                pnl = (exit_price - entry_price) * qty * position - fees
                pnl_pct = (pnl / notional_entry) * 100.0 if notional_entry else 0.0

                commission_paid += fees
                if pnl >= 0:
                    gross_profit += pnl
                    winners += 1
                else:
                    gross_loss += abs(pnl)
                    losers += 1

                equity += pnl
                equity_marks.append(equity)
                trade_returns.append(pnl_pct)
                trade_bars.append(i - entry_index)
                trades.append({
                    "entry_time": entry_time.isoformat() if entry_time is not None else None,
                    "exit_time": ts.isoformat(),
                    "side": "long" if position > 0 else "short",
                    "entry_price": round(float(entry_price), 6),
                    "exit_price": round(float(exit_price), 6),
                    "qty": int(qty),
                    "pnl": round(float(pnl), 2),
                    "pnl_pct": round(float(pnl_pct), 4),
                    "reason": exit_reason,
                    "bars_held": int(i - entry_index),
                })

                position = 0
                entry_price = 0.0
                qty = 0
                entry_index = -1
                entry_time = None
                trail_active = False
                trail_stop = None

    # Close open position at final close for metric completeness.
    if position != 0 and qty > 0:
        final_price = c[-1]
        notional_entry = qty * entry_price
        notional_exit = qty * final_price
        fees = (notional_entry + notional_exit) * (cfg.commission_pct / 100.0)
        pnl = (final_price - entry_price) * qty * position - fees
        pnl_pct = (pnl / notional_entry) * 100.0 if notional_entry else 0.0
        commission_paid += fees
        if pnl >= 0:
            gross_profit += pnl
            winners += 1
        else:
            gross_loss += abs(pnl)
            losers += 1
        equity += pnl
        equity_marks.append(equity)
        trade_returns.append(pnl_pct)
        trade_bars.append(len(df) - 1 - entry_index)
        trades.append({
            "entry_time": entry_time.isoformat() if entry_time is not None else None,
            "exit_time": idx[-1].isoformat(),
            "side": "long" if position > 0 else "short",
            "entry_price": round(float(entry_price), 6),
            "exit_price": round(float(final_price), 6),
            "qty": int(qty),
            "pnl": round(float(pnl), 2),
            "pnl_pct": round(float(pnl_pct), 4),
            "reason": "Final Close",
            "bars_held": int(len(df) - 1 - entry_index),
        })

    total_trades = winners + losers
    net_profit = equity - cfg.initial_capital
    return_pct = (net_profit / cfg.initial_capital) * 100.0 if cfg.initial_capital else 0.0
    win_rate = (winners / total_trades) * 100.0 if total_trades else 0.0
    profit_factor = gross_profit / gross_loss if gross_loss > 0 else (gross_profit if gross_profit > 0 else 0.0)

    eq = np.array(equity_marks, dtype=float)
    running_max = np.maximum.accumulate(eq)
    drawdowns = running_max - eq
    max_dd = float(drawdowns.max()) if len(drawdowns) else 0.0
    max_dd_pct = float(((drawdowns / running_max).max() * 100.0) if len(drawdowns) and np.any(running_max > 0) else 0.0)

    rets = np.array(trade_returns, dtype=float)
    sharpe = 0.0
    if len(rets) > 2 and rets.std(ddof=1) > 1e-12:
        sharpe = float(rets.mean() / rets.std(ddof=1) * math.sqrt(len(rets)))

    avg_bars = float(np.mean(trade_bars)) if trade_bars else 0.0

    score = (
        return_pct
        + (win_rate * 0.35)
        + (min(profit_factor, 10.0) * 4.0)
        + (sharpe * 2.0)
        - (max_dd_pct * 1.5)
    )
    if total_trades < 3:
        score -= 25.0

    return BacktestResult(
        params=params,
        final_equity=equity,
        net_profit=net_profit,
        return_pct=return_pct,
        total_trades=total_trades,
        winners=winners,
        losers=losers,
        win_rate_pct=win_rate,
        gross_profit=gross_profit,
        gross_loss=gross_loss,
        profit_factor=profit_factor,
        sharpe=sharpe,
        max_drawdown=max_dd,
        max_drawdown_pct=max_dd_pct,
        avg_bars_per_trade=avg_bars,
        score=score,
        trades=trades,
    )


def _finalize_result(
    params: StrategyParams,
    cfg: BacktestConfig,
    equity: float,
    equity_marks: List[float],
    winners: int,
    losers: int,
    gross_profit: float,
    gross_loss: float,
    trade_returns: List[float],
    trade_bars: List[int],
    trades: List[Dict[str, object]],
) -> BacktestResult:
    total_trades = winners + losers
    net_profit = equity - cfg.initial_capital
    return_pct = (net_profit / cfg.initial_capital) * 100.0 if cfg.initial_capital else 0.0
    win_rate = (winners / total_trades) * 100.0 if total_trades else 0.0
    profit_factor = gross_profit / gross_loss if gross_loss > 0 else (gross_profit if gross_profit > 0 else 0.0)

    eq = np.array(equity_marks, dtype=float)
    running_max = np.maximum.accumulate(eq)
    drawdowns = running_max - eq
    max_dd = float(drawdowns.max()) if len(drawdowns) else 0.0
    max_dd_pct = float(((drawdowns / running_max).max() * 100.0) if len(drawdowns) and np.any(running_max > 0) else 0.0)

    rets = np.array(trade_returns, dtype=float)
    sharpe = 0.0
    if len(rets) > 2 and rets.std(ddof=1) > 1e-12:
        sharpe = float(rets.mean() / rets.std(ddof=1) * math.sqrt(len(rets)))

    avg_bars = float(np.mean(trade_bars)) if trade_bars else 0.0
    score = (
        return_pct
        + (win_rate * 0.35)
        + (min(profit_factor, 10.0) * 4.0)
        + (sharpe * 2.0)
        - (max_dd_pct * 1.5)
    )
    if total_trades < 3:
        score -= 25.0

    return BacktestResult(
        params=params,
        final_equity=equity,
        net_profit=net_profit,
        return_pct=return_pct,
        total_trades=total_trades,
        winners=winners,
        losers=losers,
        win_rate_pct=win_rate,
        gross_profit=gross_profit,
        gross_loss=gross_loss,
        profit_factor=profit_factor,
        sharpe=sharpe,
        max_drawdown=max_dd,
        max_drawdown_pct=max_dd_pct,
        avg_bars_per_trade=avg_bars,
        score=score,
        trades=trades,
    )


def backtest_macd_sma(df: pd.DataFrame, params: StrategyParams, cfg: BacktestConfig, start_utc: datetime, end_utc: datetime) -> BacktestResult:
    df = df[(df.index >= pd.Timestamp(start_utc)) & (df.index <= pd.Timestamp(end_utc))].copy()
    if len(df) < max(100, params.macd_sma_length + params.macd_slow_length + params.macd_signal_length + 5):
        raise RuntimeError("Not enough bars for MACD/SMA backtest after date filtering.")

    fast_ma = sma(df["close"], params.macd_fast_length)
    slow_ma = sma(df["close"], params.macd_slow_length)
    veryslow_ma = sma(df["close"], params.macd_sma_length)
    macd_line = fast_ma - slow_ma
    signal_line = sma(macd_line, params.macd_signal_length)
    hist = macd_line - signal_line

    df["fast_ma"] = fast_ma
    df["slow_ma"] = slow_ma
    df["veryslow_ma"] = veryslow_ma
    df["macd_line"] = macd_line
    df["signal_line"] = signal_line
    df["hist"] = hist
    df["long_signal"] = (
        crossed_above(df["hist"], 0.0)
        & (df["macd_line"] > 0)
        & (df["close"] > df["veryslow_ma"])
    )
    df["short_signal"] = (
        crossed_below(df["hist"], 0.0)
        & (df["macd_line"] < 0)
        & (df["close"] < df["veryslow_ma"])
    )
    df = df.dropna()
    if df.empty:
        raise RuntimeError("MACD/SMA indicator warm-up removed all bars; widen date range.")

    idx = df.index
    c = df["close"].to_numpy(dtype=float)
    h = df["high"].to_numpy(dtype=float)
    l = df["low"].to_numpy(dtype=float)
    long_signal = df["long_signal"].to_numpy(dtype=bool)
    short_signal = df["short_signal"].to_numpy(dtype=bool)

    equity = cfg.initial_capital
    equity_marks: List[float] = [equity]
    position = 0
    entry_price = 0.0
    qty = 0
    entry_index = -1
    entry_time: Optional[pd.Timestamp] = None
    gross_profit = 0.0
    gross_loss = 0.0
    winners = 0
    losers = 0
    trade_returns: List[float] = []
    trade_bars: List[int] = []
    trades: List[Dict[str, object]] = []

    def close_position(i: int, exit_price: float, exit_reason: str) -> None:
        nonlocal equity, position, entry_price, qty, entry_index, entry_time, gross_profit, gross_loss, winners, losers
        notional_entry = qty * entry_price
        notional_exit = qty * exit_price
        fees = (notional_entry + notional_exit) * (cfg.commission_pct / 100.0)
        pnl = (exit_price - entry_price) * qty * position - fees
        pnl_pct = (pnl / notional_entry) * 100.0 if notional_entry else 0.0
        if pnl >= 0:
            gross_profit += pnl
            winners += 1
        else:
            gross_loss += abs(pnl)
            losers += 1
        equity += pnl
        equity_marks.append(equity)
        trade_returns.append(pnl_pct)
        trade_bars.append(i - entry_index)
        trades.append({
            "entry_time": entry_time.isoformat() if entry_time is not None else None,
            "exit_time": idx[i].isoformat(),
            "side": "long" if position > 0 else "short",
            "entry_price": round(float(entry_price), 6),
            "exit_price": round(float(exit_price), 6),
            "qty": int(qty),
            "pnl": round(float(pnl), 2),
            "pnl_pct": round(float(pnl_pct), 4),
            "reason": exit_reason,
            "bars_held": int(i - entry_index),
        })
        position = 0
        entry_price = 0.0
        qty = 0
        entry_index = -1
        entry_time = None

    close_min = cfg.market_close_utc_hour * 60 + cfg.market_close_utc_minute
    preclose_start = close_min - cfg.close_before_minutes

    for i in range(1, len(df)):
        if position != 0:
            # Force close before market end to avoid overnight holds.
            bar_min = idx[i].hour * 60 + idx[i].minute
            if preclose_start <= bar_min < close_min:
                close_position(i, c[i], "MACD Market Close")
                continue

            exit_price: Optional[float] = None
            exit_reason: Optional[str] = None
            if position > 0:
                if l[i] <= entry_price * (1 - params.forced_stop_loss_pct / 100.0):
                    exit_price = entry_price * (1 - params.forced_stop_loss_pct / 100.0)
                    exit_reason = "MACD Forced SL Long"
                elif h[i] >= entry_price * (1 + params.forced_take_profit_pct / 100.0):
                    exit_price = entry_price * (1 + params.forced_take_profit_pct / 100.0)
                    exit_reason = "MACD Forced TP Long"
                elif c[i] <= entry_price * (1 - params.fixed_stop_loss_pct / 100.0):
                    exit_price = c[i]
                    exit_reason = "MACD Fixed Stop Loss Long"
                elif c[i] >= entry_price * (1 + params.fixed_take_profit_pct / 100.0):
                    exit_price = c[i]
                    exit_reason = "MACD Fixed Take Profit Long"
                elif short_signal[i]:
                    exit_price = c[i]
                    exit_reason = "MACD Opposite Short Signal"
            else:
                if h[i] >= entry_price * (1 + params.forced_stop_loss_pct / 100.0):
                    exit_price = entry_price * (1 + params.forced_stop_loss_pct / 100.0)
                    exit_reason = "MACD Forced SL Short"
                elif l[i] <= entry_price * (1 - params.forced_take_profit_pct / 100.0):
                    exit_price = entry_price * (1 - params.forced_take_profit_pct / 100.0)
                    exit_reason = "MACD Forced TP Short"
                elif c[i] >= entry_price * (1 + params.fixed_stop_loss_pct / 100.0):
                    exit_price = c[i]
                    exit_reason = "MACD Fixed Stop Loss Short"
                elif c[i] <= entry_price * (1 - params.fixed_take_profit_pct / 100.0):
                    exit_price = c[i]
                    exit_reason = "MACD Fixed Take Profit Short"
                elif long_signal[i]:
                    exit_price = c[i]
                    exit_reason = "MACD Opposite Long Signal"
            if exit_price is not None and exit_reason is not None:
                close_position(i, float(exit_price), exit_reason)
                continue

        if position == 0:
            entry_side = 0
            if params.trade_direction in ("Both", "Long Only") and long_signal[i]:
                entry_side = 1
            elif params.trade_direction in ("Both", "Short Only") and short_signal[i]:
                entry_side = -1
            if entry_side:
                q = math.floor(cfg.order_size_usd / c[i])
                if q > 0:
                    position = entry_side
                    entry_price = c[i]
                    qty = q
                    entry_index = i
                    entry_time = idx[i]

    if position != 0 and qty > 0:
        close_position(len(df) - 1, float(c[-1]), "Final Close")

    return _finalize_result(
        params=params,
        cfg=cfg,
        equity=equity,
        equity_marks=equity_marks,
        winners=winners,
        losers=losers,
        gross_profit=gross_profit,
        gross_loss=gross_loss,
        trade_returns=trade_returns,
        trade_bars=trade_bars,
        trades=trades,
    )


def run_strategy_backtest(strategy: str, df: pd.DataFrame, params: StrategyParams, cfg: BacktestConfig, start_utc: datetime, end_utc: datetime) -> BacktestResult:
    if strategy == "macd_sma":
        return backtest_macd_sma(df, params, cfg, start_utc, end_utc)
    if NUMBA_AVAILABLE:
        return backtest_fast(df, params, cfg, start_utc, end_utc)
    return backtest(df, params, cfg, start_utc, end_utc)


def parse_range(spec: str, is_int: bool) -> List[float]:
    parts = [x.strip() for x in spec.split(":")]
    if len(parts) != 3:
        raise ValueError(f"Range must be min:max:step, got '{spec}'")
    start, end, step = map(float, parts)
    if step <= 0:
        raise ValueError("Step must be > 0")
    values: List[float] = []
    x = start
    while x <= end + 1e-12:
        values.append(round(x if not is_int else int(round(x)), 10))
        x += step
    if is_int:
        values = sorted(set(int(v) for v in values))
    return values


def sample_params(
    rng: random.Random,
    base: StrategyParams,
    ranges: Dict[str, List[float]],
    trade_direction: str,
) -> StrategyParams:
    return StrategyParams(
        trade_direction=trade_direction,
        inner_kc_length=int(rng.choice(ranges["inner_kc_length"])),
        inner_kc_mult=float(rng.choice(ranges["inner_kc_mult"])),
        outer_kc_length=int(rng.choice(ranges["outer_kc_length"])),
        outer_kc_mult=float(rng.choice(ranges["outer_kc_mult"])),
        fixed_stop_loss_pct=float(rng.choice(ranges["fixed_stop_loss_pct"])),
        fixed_take_profit_pct=float(rng.choice(ranges["fixed_take_profit_pct"])),
        forced_stop_loss_pct=float(rng.choice(ranges["forced_stop_loss_pct"])),
        forced_take_profit_pct=float(rng.choice(ranges["forced_take_profit_pct"])),
        trailing_offset_ticks=int(rng.choice(ranges["trailing_offset_ticks"])),
        tick_size=base.tick_size,
        trailing_offset_pct=float(rng.choice(ranges.get("trailing_offset_pct", [base.trailing_offset_pct]))),
        macd_fast_length=int(rng.choice(ranges.get("macd_fast_length", [base.macd_fast_length]))),
        macd_slow_length=int(rng.choice(ranges.get("macd_slow_length", [base.macd_slow_length]))),
        macd_signal_length=int(rng.choice(ranges.get("macd_signal_length", [base.macd_signal_length]))),
        macd_sma_length=int(rng.choice(ranges.get("macd_sma_length", [base.macd_sma_length]))),
        max_intraday_loss_pct=float(rng.choice(ranges.get("max_intraday_loss_pct", [base.max_intraday_loss_pct]))),
    )


def result_row(res: BacktestResult) -> Dict[str, object]:
    row = {
        "score": round(res.score, 6),
        "net_profit": round(res.net_profit, 2),
        "return_pct": round(res.return_pct, 2),
        "total_trades": res.total_trades,
        "win_rate_pct": round(res.win_rate_pct, 2),
        "profit_factor": round(res.profit_factor, 4),
        "sharpe": round(res.sharpe, 4),
        "max_drawdown": round(res.max_drawdown, 2),
        "max_drawdown_pct": round(res.max_drawdown_pct, 2),
    }
    row.update(asdict(res.params))
    return row

def result_row_for_timeframe(timeframe: str, res: BacktestResult) -> Dict[str, object]:
    row = {"timeframe": timeframe}
    row.update(result_row(res))
    return row


def split_train_test_bars(df: pd.DataFrame, train_ratio: float) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if df.empty:
        return df, df.iloc[0:0]
    train_ratio = min(0.95, max(0.2, float(train_ratio)))
    split_at = int(len(df) * train_ratio)
    split_at = min(max(split_at, 1), max(1, len(df) - 1))
    return df.iloc[:split_at].copy(), df.iloc[split_at:].copy()


def result_or_error(strategy: str, df: pd.DataFrame, params: StrategyParams, cfg: BacktestConfig, start_utc: datetime, end_utc: datetime) -> Tuple[Optional[BacktestResult], Optional[str]]:
    try:
        return run_strategy_backtest(strategy, df, params, cfg, start_utc, end_utc), None
    except Exception as exc:
        return None, str(exc)


def validation_status(row: Optional[Dict[str, object]], args) -> Dict[str, object]:
    checks = {
        "min_net_profit": {"threshold": float(args.validation_min_net_profit), "actual": None, "pass": False},
        "min_trades": {"threshold": int(args.validation_min_trades), "actual": None, "pass": False},
        "min_win_rate_pct": {"threshold": float(args.validation_min_win_rate), "actual": None, "pass": False},
        "min_profit_factor": {"threshold": float(args.validation_min_profit_factor), "actual": None, "pass": False},
        "max_drawdown_pct": {"threshold": float(args.validation_max_drawdown_pct), "actual": None, "pass": False},
    }
    if not isinstance(row, dict):
        return {"passed": False, "checks": checks, "reason": "missing_validation_result"}

    def f(key: str) -> Optional[float]:
        try:
            value = row.get(key)
            if value is None:
                return None
            return float(value)
        except Exception:
            return None

    values = {
        "min_net_profit": f("net_profit"),
        "min_trades": f("total_trades"),
        "min_win_rate_pct": f("win_rate_pct"),
        "min_profit_factor": f("profit_factor"),
        "max_drawdown_pct": f("max_drawdown_pct"),
    }
    checks["min_net_profit"]["actual"] = values["min_net_profit"]
    checks["min_trades"]["actual"] = values["min_trades"]
    checks["min_win_rate_pct"]["actual"] = values["min_win_rate_pct"]
    checks["min_profit_factor"]["actual"] = values["min_profit_factor"]
    checks["max_drawdown_pct"]["actual"] = values["max_drawdown_pct"]
    checks["min_net_profit"]["pass"] = values["min_net_profit"] is not None and values["min_net_profit"] >= float(args.validation_min_net_profit)
    checks["min_trades"]["pass"] = values["min_trades"] is not None and values["min_trades"] >= int(args.validation_min_trades)
    checks["min_win_rate_pct"]["pass"] = values["min_win_rate_pct"] is not None and values["min_win_rate_pct"] >= float(args.validation_min_win_rate)
    checks["min_profit_factor"]["pass"] = values["min_profit_factor"] is not None and values["min_profit_factor"] >= float(args.validation_min_profit_factor)
    checks["max_drawdown_pct"]["pass"] = values["max_drawdown_pct"] is not None and values["max_drawdown_pct"] <= float(args.validation_max_drawdown_pct)
    failed = [name for name, check in checks.items() if not check["pass"]]
    return {"passed": not failed, "checks": checks, "failed_checks": failed}


def init_backtest_worker(df: pd.DataFrame, cfg: BacktestConfig, start_utc: datetime, end_utc: datetime, strategy: str = "keltner") -> None:
    global _WORKER_DF, _WORKER_CFG, _WORKER_START, _WORKER_END, _WORKER_STRATEGY
    _WORKER_DF = df
    _WORKER_CFG = cfg
    _WORKER_START = start_utc
    _WORKER_END = end_utc
    _WORKER_STRATEGY = strategy


def safe_backtest_worker(params: StrategyParams) -> Optional[BacktestResult]:
    try:
        if _WORKER_DF is None or _WORKER_CFG is None or _WORKER_START is None or _WORKER_END is None:
            raise RuntimeError("Backtest worker was not initialized.")
        return run_strategy_backtest(_WORKER_STRATEGY, _WORKER_DF, params, _WORKER_CFG, _WORKER_START, _WORKER_END)
    except Exception:
        return None


def resolve_optimizer_jobs(requested: int) -> int:
    if requested > 0:
        return requested
    cpu_count = os.cpu_count() or 1
    return max(1, cpu_count - 1)


def describe_accelerator(requested: str) -> Dict[str, object]:
    requested = (requested or "auto").strip().lower()
    if requested not in {"auto", "cpu", "gpu"}:
        requested = "auto"
    info: Dict[str, object] = {
        "requested": requested,
        "active": "cpu",
        "gpu_available": False,
        "backend": None,
        "message": "CPU execution selected.",
    }
    if requested == "cpu":
        return info

    probes: List[str] = []
    # NOTE: the backtest is a sequential, path-dependent state machine (entries,
    # intrabar trailing stops, forced SL/TP per bar). That does not map onto GPU
    # SIMD without a full vectorized rewrite, so even when a GPU is present the
    # simulation runs on the CPU and trials are parallelized across CPU processes
    # (see --jobs). Detection below reports the GPU honestly rather than pretending
    # to offload work to it.
    try:
        import torch_directml  # type: ignore

        device = torch_directml.device()
        try:
            device_name = torch_directml.device_name(0)
        except Exception:
            device_name = str(device)
        info["gpu_available"] = True
        info["backend"] = "torch-directml"
        info["device_name"] = device_name
        info["message"] = (
            f"GPU detected via DirectML ({device_name}). The OHLC simulator is sequential and runs on CPU; "
            "trials are parallelized across CPU cores (set optimizer jobs to 0 for auto)."
        )
        return info
    except Exception as exc:
        probes.append(f"torch-directml unavailable: {exc}")

    try:
        import pyopencl as cl  # type: ignore

        platforms = cl.get_platforms()
        devices = [dev.name for platform in platforms for dev in platform.get_devices(device_type=cl.device_type.GPU)]
        if devices:
            info["gpu_available"] = True
            info["backend"] = "opencl"
            info["device_name"] = devices[0]
            info["message"] = (
                f"GPU detected via OpenCL ({devices[0]}). The OHLC simulator is sequential and runs on CPU; "
                "trials are parallelized across CPU cores (set optimizer jobs to 0 for auto)."
            )
            return info
    except Exception as exc:
        probes.append(f"pyopencl unavailable: {exc}")

    if requested == "gpu":
        info["message"] = "GPU was requested, but no supported Python GPU backend was detected; running on CPU."
    else:
        info["message"] = "No supported Python GPU backend detected; running on CPU."
    info["probes"] = probes[-3:]
    return info


def suggest_params_tpe(
    trial,
    base: StrategyParams,
    ranges: Dict[str, List[float]],
    trade_direction: str,
) -> StrategyParams:
    return StrategyParams(
        trade_direction=trade_direction,
        inner_kc_length=int(trial.suggest_categorical("inner_kc_length", ranges["inner_kc_length"])),
        inner_kc_mult=float(trial.suggest_categorical("inner_kc_mult", ranges["inner_kc_mult"])),
        outer_kc_length=int(trial.suggest_categorical("outer_kc_length", ranges["outer_kc_length"])),
        outer_kc_mult=float(trial.suggest_categorical("outer_kc_mult", ranges["outer_kc_mult"])),
        fixed_stop_loss_pct=float(trial.suggest_categorical("fixed_stop_loss_pct", ranges["fixed_stop_loss_pct"])),
        fixed_take_profit_pct=float(trial.suggest_categorical("fixed_take_profit_pct", ranges["fixed_take_profit_pct"])),
        forced_stop_loss_pct=float(trial.suggest_categorical("forced_stop_loss_pct", ranges["forced_stop_loss_pct"])),
        forced_take_profit_pct=float(trial.suggest_categorical("forced_take_profit_pct", ranges["forced_take_profit_pct"])),
        trailing_offset_ticks=int(trial.suggest_categorical("trailing_offset_ticks", ranges["trailing_offset_ticks"])),
        tick_size=base.tick_size,
        trailing_offset_pct=float(trial.suggest_categorical("trailing_offset_pct", ranges.get("trailing_offset_pct", [base.trailing_offset_pct]))),
        macd_fast_length=int(trial.suggest_categorical("macd_fast_length", ranges.get("macd_fast_length", [base.macd_fast_length]))),
        macd_slow_length=int(trial.suggest_categorical("macd_slow_length", ranges.get("macd_slow_length", [base.macd_slow_length]))),
        macd_signal_length=int(trial.suggest_categorical("macd_signal_length", ranges.get("macd_signal_length", [base.macd_signal_length]))),
        macd_sma_length=int(trial.suggest_categorical("macd_sma_length", ranges.get("macd_sma_length", [base.macd_sma_length]))),
        max_intraday_loss_pct=float(trial.suggest_categorical("max_intraday_loss_pct", ranges.get("max_intraday_loss_pct", [base.max_intraday_loss_pct]))),
    )


def run_tpe_trials(
    tf_label: str,
    tf_bars: pd.DataFrame,
    cfg: BacktestConfig,
    start_utc: datetime,
    end_utc: datetime,
    params_template: StrategyParams,
    ranges: Dict[str, List[float]],
    trade_direction: str,
    strategy: str,
    trials: int,
    seed: int,
    all_results: List[Tuple[str, BacktestResult]],
) -> int:
    try:
        import optuna
    except Exception as exc:
        raise RuntimeError(
            "Optuna TPE selected but optuna is not installed. Run `pip install -r requirements.txt` "
            "on the machine running the optimizer."
        ) from exc

    optuna.logging.set_verbosity(optuna.logging.WARNING)
    sampler = optuna.samplers.TPESampler(seed=seed)
    study = optuna.create_study(direction="maximize", sampler=sampler)
    completed = 0

    def objective(trial) -> float:
        nonlocal completed
        params = suggest_params_tpe(trial, params_template, ranges, trade_direction)
        try:
            res = run_strategy_backtest(strategy, tf_bars, params, cfg, start_utc, end_utc)
        except Exception:
            return -1_000_000_000.0
        all_results.append((tf_label, res))
        completed += 1
        if completed % 25 == 0:
            best_tf, best_res = max(all_results, key=lambda item: item[1].score)
            print(f"[{tf_label}] TPE trial {completed:4d}/{trials}: global best={best_tf} "
                  f"score={best_res.score:.4f}, net={best_res.net_profit:.2f}, "
                  f"PF={best_res.profit_factor:.3f}, trades={best_res.total_trades}")
        return float(res.score)

    study.optimize(objective, n_trials=max(1, trials), n_jobs=1, show_progress_bar=False)
    if completed and completed % 25 != 0:
        best_tf, best_res = max(all_results, key=lambda item: item[1].score)
        print(f"[{tf_label}] TPE trial {completed:4d}/{trials}: global best={best_tf} "
              f"score={best_res.score:.4f}, net={best_res.net_profit:.2f}, "
              f"PF={best_res.profit_factor:.3f}, trades={best_res.total_trades}")
    return completed


def main() -> None:
    parser = argparse.ArgumentParser(description="Optimize project Pine strategy parameters locally.")
    parser.add_argument("--strategy", type=str, default="keltner", choices=sorted(SUPPORTED_STRATEGIES))
    parser.add_argument("--optimizer-engine", type=str, default="random", choices=["random", "tpe"])
    parser.add_argument("--accelerator", type=str, default="auto", choices=["auto", "cpu", "gpu"])
    parser.add_argument("--pine", type=Path, default=None)
    parser.add_argument("--reference-xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--bars-csv", type=Path, default=None, help="Optional CSV with timestamp,open,high,low,close[,volume]")
    parser.add_argument("--alpaca-user", type=str, default=None, help="Username from local DB to use Alpaca keys")
    parser.add_argument("--feed", type=str, default="iex", help="Alpaca feed (iex/sip). For free plans use iex.")
    parser.add_argument("--symbol", type=str, default=None)
    parser.add_argument("--timeframe", type=str, default=None)
    parser.add_argument("--timeframes", type=str, default=None, help="Comma-separated sweep, e.g. 5Min,10Min,1Hour,2Day")
    parser.add_argument("--session", type=str, default="regular", choices=["regular", "extended", "all"])
    parser.add_argument("--start", type=str, default=None, help="UTC ISO start override, e.g. 2023-01-03T00:00:00Z")
    parser.add_argument("--end", type=str, default=None, help="UTC ISO end override")
    parser.add_argument("--timezone", type=str, default="Europe/Bucharest")
    parser.add_argument("--initial-capital", type=float, default=None)
    parser.add_argument("--order-size", type=float, default=None)
    parser.add_argument("--commission-pct", type=float, default=None)
    parser.add_argument("--tick-size", type=float, default=None)
    parser.add_argument("--trials", type=int, default=250)
    parser.add_argument("--jobs", type=int, default=1, help="Parallel optimizer processes. Use 0 for auto cpu_count-1.")
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--top-k", type=int, default=20)
    parser.add_argument("--trade-direction", type=str, default="Both", choices=["Both", "Long Only", "Short Only"])
    parser.add_argument("--validation-enabled", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--validation-train-ratio", type=float, default=0.70)
    parser.add_argument("--validation-min-trades", type=int, default=30)
    parser.add_argument("--validation-min-win-rate", type=float, default=55.0)
    parser.add_argument("--validation-min-profit-factor", type=float, default=1.3)
    parser.add_argument("--validation-max-drawdown-pct", type=float, default=8.0)
    parser.add_argument("--validation-min-net-profit", type=float, default=0.0)

    parser.add_argument("--inner-len-range", type=str, default="8:40:1")
    parser.add_argument("--inner-mult-range", type=str, default="0.6:1.8:0.1")
    parser.add_argument("--outer-len-range", type=str, default="8:40:1")
    # Outer multiplier is integer in the current Pine input definition (`input(2, ...)`).
    parser.add_argument("--outer-mult-range", type=str, default="1:4:1")
    parser.add_argument("--fixed-sl-range", type=str, default="1.0:5.0:0.1")
    parser.add_argument("--fixed-tp-range", type=str, default="0.8:4.0:0.1")
    parser.add_argument("--forced-sl-range", type=str, default="3.0:10.0:0.2")
    parser.add_argument("--forced-tp-range", type=str, default="3.0:10.0:0.2")
    # Trailing offset is currently hardcoded in Pine (`trail_offset = 4`), so keep fixed by default.
    parser.add_argument("--trail-offset-range", type=str, default="4:4:1")
    parser.add_argument("--trail-pct-range", type=str, default="0:0:1")
    parser.add_argument("--macd-fast-range", type=str, default="8:20:1")
    parser.add_argument("--macd-slow-range", type=str, default="20:40:1")
    parser.add_argument("--macd-signal-range", type=str, default="5:15:1")
    parser.add_argument("--macd-sma-range", type=str, default="100:250:10")
    parser.add_argument("--max-intraday-loss-range", type=str, default="50:50:1")

    parser.add_argument("--report-json", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--top-csv", type=Path, default=DEFAULT_TOP_CSV)

    args = parser.parse_args()

    strategy_name = args.strategy.strip().lower()
    if strategy_name not in SUPPORTED_STRATEGIES:
        raise SystemExit(f"Unsupported strategy '{args.strategy}'. Supported: {', '.join(sorted(SUPPORTED_STRATEGIES))}")

    accelerator_info = describe_accelerator(args.accelerator)
    print(f"Strategy: {strategy_name}")
    print(f"Optimizer engine: {args.optimizer_engine}")
    print(f"Accelerator: requested={accelerator_info['requested']} active={accelerator_info['active']} - {accelerator_info['message']}")

    pine_path = args.pine or (DEFAULT_MACD_SMA_PINE if strategy_name == "macd_sma" else DEFAULT_KELTNER_PINE)
    if not pine_path.exists():
        raise SystemExit(f"Pine file not found: {pine_path}")

    pine_defaults = parse_pine_defaults(pine_path)

    ref_data = {"properties": {}}
    if args.reference_xlsx and args.reference_xlsx.exists():
        ref_data = read_reference_xlsx(args.reference_xlsx)

    ref_props = ref_data.get("properties", {})
    base_params, cfg, symbol_from_ref, tf_from_ref, start_utc, end_utc, raw_symbol = derive_reference_settings(
        pine_defaults,
        ref_props,
        args.timezone,
    )

    symbol = (args.symbol or symbol_from_ref).split(":")[-1].upper()
    timeframe = args.timeframe or tf_from_ref
    timeframes = parse_timeframes_arg(args.timeframes, timeframe)
    fetch_timeframe = choose_fetch_timeframe(timeframes)

    if args.start:
        start_utc = pd.Timestamp(args.start, tz="UTC").to_pydatetime()
    if args.end:
        end_utc = pd.Timestamp(args.end, tz="UTC").to_pydatetime()

    if args.initial_capital is not None:
        cfg.initial_capital = float(args.initial_capital)
    if args.order_size is not None:
        cfg.order_size_usd = float(args.order_size)
    if args.commission_pct is not None:
        cfg.commission_pct = float(args.commission_pct)
    if args.tick_size is not None:
        base_params.tick_size = float(args.tick_size)
    base_params.macd_fast_length = int(pine_defaults.get("macd_fast_length", 12))
    base_params.macd_slow_length = int(pine_defaults.get("macd_slow_length", 26))
    base_params.macd_signal_length = int(pine_defaults.get("macd_signal_length", 9))
    base_params.macd_sma_length = int(pine_defaults.get("macd_sma_length", 200))
    base_params.max_intraday_loss_pct = float(pine_defaults.get("max_intraday_loss_pct", 50.0))

    if args.bars_csv:
        bars = load_bars_csv(args.bars_csv)
    else:
        bars = fetch_bars_alpaca(
            symbol=symbol,
            timeframe=fetch_timeframe,
            start_utc=start_utc,
            end_utc=end_utc,
            username=args.alpaca_user,
            feed=args.feed,
        )

    ranges = {
        "inner_kc_length": parse_range(args.inner_len_range, is_int=True),
        "inner_kc_mult": parse_range(args.inner_mult_range, is_int=False),
        "outer_kc_length": parse_range(args.outer_len_range, is_int=True),
        "outer_kc_mult": parse_range(args.outer_mult_range, is_int=False),
        "fixed_stop_loss_pct": parse_range(args.fixed_sl_range, is_int=False),
        "fixed_take_profit_pct": parse_range(args.fixed_tp_range, is_int=False),
        "forced_stop_loss_pct": parse_range(args.forced_sl_range, is_int=False),
        "forced_take_profit_pct": parse_range(args.forced_tp_range, is_int=False),
        "trailing_offset_ticks": parse_range(args.trail_offset_range, is_int=True),
        "trailing_offset_pct": parse_range(args.trail_pct_range, is_int=False),
        "macd_fast_length": parse_range(args.macd_fast_range, is_int=True),
        "macd_slow_length": parse_range(args.macd_slow_range, is_int=True),
        "macd_signal_length": parse_range(args.macd_signal_range, is_int=True),
        "macd_sma_length": parse_range(args.macd_sma_range, is_int=True),
        "max_intraday_loss_pct": parse_range(args.max_intraday_loss_range, is_int=False),
    }
    if strategy_name == "keltner":
        ranges["macd_fast_length"] = [base_params.macd_fast_length]
        ranges["macd_slow_length"] = [base_params.macd_slow_length]
        ranges["macd_signal_length"] = [base_params.macd_signal_length]
        ranges["macd_sma_length"] = [base_params.macd_sma_length]
        ranges["max_intraday_loss_pct"] = [base_params.max_intraday_loss_pct]
    elif strategy_name == "macd_sma":
        ranges["inner_kc_length"] = [base_params.inner_kc_length]
        ranges["inner_kc_mult"] = [base_params.inner_kc_mult]
        ranges["outer_kc_length"] = [base_params.outer_kc_length]
        ranges["outer_kc_mult"] = [base_params.outer_kc_mult]
        ranges["trailing_offset_ticks"] = [base_params.trailing_offset_ticks]
        ranges["trailing_offset_pct"] = [base_params.trailing_offset_pct]

    all_results: List[Tuple[str, BacktestResult]] = []
    reference_timeframe = None
    reference_result = None
    bars_count_by_timeframe: Dict[str, int] = {}
    train_bars_count_by_timeframe: Dict[str, int] = {}
    test_bars_count_by_timeframe: Dict[str, int] = {}
    bars_by_timeframe: Dict[str, pd.DataFrame] = {}
    train_ranges_by_timeframe: Dict[str, Dict[str, Optional[str]]] = {}
    test_ranges_by_timeframe: Dict[str, Dict[str, Optional[str]]] = {}
    optimizer_jobs = resolve_optimizer_jobs(args.jobs)
    if args.optimizer_engine == "tpe":
        print("Optuna TPE runs sequentially in this implementation; CPU Jobs is used by the random engine.")
    elif optimizer_jobs > 1:
        print(f"Optimizer parallelism: {optimizer_jobs} processes")

    for tf_index, tf_label in enumerate(timeframes):
        tf_bars = resample_bars(bars, tf_label)
        tf_bars = filter_session(tf_bars, args.session)
        bars_count_by_timeframe[tf_label] = int(len(tf_bars))
        bars_by_timeframe[tf_label] = tf_bars
        if tf_bars.empty:
            print(f"Skipping {tf_label}: no bars after session filter '{args.session}'.")
            continue
        train_bars, test_bars = split_train_test_bars(tf_bars, args.validation_train_ratio) if args.validation_enabled else (tf_bars, tf_bars.iloc[0:0])
        optimize_bars = train_bars if args.validation_enabled else tf_bars
        train_bars_count_by_timeframe[tf_label] = int(len(train_bars))
        test_bars_count_by_timeframe[tf_label] = int(len(test_bars))
        train_ranges_by_timeframe[tf_label] = {
            "start": train_bars.index[0].isoformat() if not train_bars.empty else None,
            "end": train_bars.index[-1].isoformat() if not train_bars.empty else None,
        }
        test_ranges_by_timeframe[tf_label] = {
            "start": test_bars.index[0].isoformat() if not test_bars.empty else None,
            "end": test_bars.index[-1].isoformat() if not test_bars.empty else None,
        }

        params_template = deepcopy(base_params)
        if all(len(values) == 1 for values in ranges.values()):
            params_template = StrategyParams(
                trade_direction=args.trade_direction,
                inner_kc_length=int(ranges["inner_kc_length"][0]),
                inner_kc_mult=float(ranges["inner_kc_mult"][0]),
                outer_kc_length=int(ranges["outer_kc_length"][0]),
                outer_kc_mult=float(ranges["outer_kc_mult"][0]),
                fixed_stop_loss_pct=float(ranges["fixed_stop_loss_pct"][0]),
                fixed_take_profit_pct=float(ranges["fixed_take_profit_pct"][0]),
                forced_stop_loss_pct=float(ranges["forced_stop_loss_pct"][0]),
                forced_take_profit_pct=float(ranges["forced_take_profit_pct"][0]),
                trailing_offset_ticks=int(ranges["trailing_offset_ticks"][0]),
                tick_size=params_template.tick_size,
                macd_fast_length=int(ranges["macd_fast_length"][0]),
                macd_slow_length=int(ranges["macd_slow_length"][0]),
                macd_signal_length=int(ranges["macd_signal_length"][0]),
                macd_sma_length=int(ranges["macd_sma_length"][0]),
                max_intraday_loss_pct=float(ranges["max_intraday_loss_pct"][0]),
            )

        rng = random.Random(args.seed + tf_index)
        seen = set()
        params_template.trade_direction = args.trade_direction
        try:
            base_result = run_strategy_backtest(strategy_name, optimize_bars, params_template, cfg, start_utc, end_utc)
        except Exception as exc:
            print(f"Skipping {tf_label}: {exc}")
            continue
        all_results.append((tf_label, base_result))
        if reference_result is None:
            reference_result = base_result
            reference_timeframe = tf_label
        seen.add(tuple(asdict(params_template).items()))

        print(f"[{tf_label}] Reference/backbone result: net={base_result.net_profit:.2f} USD | "
              f"PF={base_result.profit_factor:.3f} | trades={base_result.total_trades} | "
              f"win={base_result.win_rate_pct:.2f}%")

        has_search_space = any(len(values) > 1 for values in ranges.values())
        if args.optimizer_engine == "tpe" and has_search_space:
            run_tpe_trials(
                tf_label=tf_label,
                tf_bars=optimize_bars,
                cfg=cfg,
                start_utc=start_utc,
                end_utc=end_utc,
                params_template=params_template,
                ranges=ranges,
                trade_direction=args.trade_direction,
                strategy=strategy_name,
                trials=args.trials,
                seed=args.seed + tf_index,
                all_results=all_results,
            )
            continue

        sampled_params: List[StrategyParams] = []
        for _ in range(args.trials):
            p = sample_params(rng, params_template, ranges, args.trade_direction)
            key = tuple(asdict(p).items())
            if key in seen:
                continue
            seen.add(key)
            sampled_params.append(p)

        if optimizer_jobs <= 1 or len(sampled_params) <= 1:
            result_iter = (safe_backtest_worker(p) for p in sampled_params)
            init_backtest_worker(optimize_bars, cfg, start_utc, end_utc, strategy_name)
            for i, res in enumerate(result_iter, start=1):
                if res is not None:
                    all_results.append((tf_label, res))
                if i % 25 == 0:
                    best_tf, best_res = max(all_results, key=lambda item: item[1].score)
                    print(f"[{tf_label}] Trial {i:4d}/{len(sampled_params)}: global best={best_tf} "
                          f"score={best_res.score:.4f}, net={best_res.net_profit:.2f}, "
                          f"PF={best_res.profit_factor:.3f}, trades={best_res.total_trades}")
        else:
            chunksize = max(1, len(sampled_params) // (optimizer_jobs * 8))
            with ProcessPoolExecutor(
                max_workers=optimizer_jobs,
                initializer=init_backtest_worker,
                initargs=(optimize_bars, cfg, start_utc, end_utc, strategy_name),
            ) as executor:
                for i, res in enumerate(executor.map(safe_backtest_worker, sampled_params, chunksize=chunksize), start=1):
                    if res is not None:
                        all_results.append((tf_label, res))
                    if i % 25 == 0:
                        best_tf, best_res = max(all_results, key=lambda item: item[1].score)
                        print(f"[{tf_label}] Trial {i:4d}/{len(sampled_params)}: global best={best_tf} "
                              f"score={best_res.score:.4f}, net={best_res.net_profit:.2f}, "
                              f"PF={best_res.profit_factor:.3f}, trades={best_res.total_trades}")

        if sampled_params:
            best_tf, best_res = max(all_results, key=lambda item: item[1].score)
            if len(sampled_params) % 25 != 0:
                best_tf, best_res = max(all_results, key=lambda item: item[1].score)
                print(f"[{tf_label}] Trial {len(sampled_params):4d}/{len(sampled_params)}: global best={best_tf} "
                      f"score={best_res.score:.4f}, net={best_res.net_profit:.2f}, "
                      f"PF={best_res.profit_factor:.3f}, trades={best_res.total_trades}")

    if not all_results:
        raise SystemExit(f"No bars left after applying session filter '{args.session}' for any timeframe.")

    all_results.sort(key=lambda item: item[1].score, reverse=True)
    top = all_results[: max(1, args.top_k)]

    args.top_csv.parent.mkdir(parents=True, exist_ok=True)
    with args.top_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(result_row_for_timeframe(top[0][0], top[0][1]).keys()))
        writer.writeheader()
        for tf_label, res in top:
            writer.writerow(result_row_for_timeframe(tf_label, res))

    best_timeframe, best_train = top[0]
    best_params = best_train.params
    best_bars = bars_by_timeframe.get(best_timeframe)
    best_full = best_train
    full_error = None
    if best_bars is not None:
        best_full, full_error = result_or_error(strategy_name, best_bars, best_params, cfg, start_utc, end_utc)
        if best_full is None:
            best_full = best_train

    train_result = best_train
    test_result = None
    test_error = None
    if args.validation_enabled and best_bars is not None:
        train_bars, test_bars = split_train_test_bars(best_bars, args.validation_train_ratio)
        train_result, train_error = result_or_error(strategy_name, train_bars, best_params, cfg, start_utc, end_utc)
        if train_result is None:
            train_result = best_train
        test_result, test_error = result_or_error(strategy_name, test_bars, best_params, cfg, start_utc, end_utc)
    else:
        train_error = None

    best = best_full
    train_row = result_row_for_timeframe(best_timeframe, train_result) if train_result is not None else None
    test_row = result_row_for_timeframe(best_timeframe, test_result) if test_result is not None else None
    full_row = result_row_for_timeframe(best_timeframe, best_full)
    validation = {
        "enabled": bool(args.validation_enabled),
        "method": "single_train_test_split",
        "train_ratio": float(args.validation_train_ratio),
        "train_date_range_utc": train_ranges_by_timeframe.get(best_timeframe),
        "test_date_range_utc": test_ranges_by_timeframe.get(best_timeframe),
        "thresholds": {
            "min_net_profit": float(args.validation_min_net_profit),
            "min_trades": int(args.validation_min_trades),
            "min_win_rate_pct": float(args.validation_min_win_rate),
            "min_profit_factor": float(args.validation_min_profit_factor),
            "max_drawdown_pct": float(args.validation_max_drawdown_pct),
        },
        "optimization_result": train_row,
        "test_result": test_row,
        "full_result": full_row,
        "status": validation_status(test_row, args) if args.validation_enabled else {"passed": True, "checks": {}, "reason": "disabled"},
        "errors": {"train": train_error, "test": test_error, "full": full_error},
    }
    report = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "strategy": strategy_name,
        "optimizer_engine": args.optimizer_engine,
        "accelerator": accelerator_info,
        "symbol_input": raw_symbol,
        "symbol_used": symbol,
        "timeframe": best_timeframe if len(timeframes) == 1 else "sweep",
        "timeframes": timeframes,
        "best_timeframe": best_timeframe,
        "fetch_timeframe": fetch_timeframe,
        "date_range_utc": {"start": start_utc.isoformat(), "end": end_utc.isoformat()},
        "bars_count": int(bars_count_by_timeframe.get(best_timeframe, 0)),
        "bars_count_by_timeframe": bars_count_by_timeframe,
        "train_bars_count_by_timeframe": train_bars_count_by_timeframe,
        "test_bars_count_by_timeframe": test_bars_count_by_timeframe,
        "session_filter": args.session,
        "config": asdict(cfg),
        "reference_properties": {k: str(v) for k, v in ref_props.items()},
        "reference_metrics": ref_data,
        "reference_result": result_row_for_timeframe(reference_timeframe or best_timeframe, reference_result or best),
        "best_result": full_row,
        "validation": validation,
        "best_trades": best.trades,
        "top_results": [result_row_for_timeframe(tf_label, res) for tf_label, res in top],
    }

    args.report_json.parent.mkdir(parents=True, exist_ok=True)
    args.report_json.write_text(json.dumps(report, indent=2), encoding="utf-8")

    print("\n=== Best combination ===")
    print(f"Timeframe: {best_timeframe}")
    print(f"Score: {best.score:.4f}")
    print(f"Net profit: {best.net_profit:.2f} USD ({best.return_pct:.2f}%)")
    print(f"Profit factor: {best.profit_factor:.3f} | Sharpe: {best.sharpe:.3f}")
    print(f"Trades: {best.total_trades} | Win rate: {best.win_rate_pct:.2f}% | Max DD: {best.max_drawdown_pct:.2f}%")
    if args.validation_enabled:
        print(f"Out-of-sample validation: {'PASS' if validation['status'].get('passed') else 'FAIL'}")
    print("Params:")
    for k, v in asdict(best.params).items():
        print(f"  - {k}: {v}")

    print(f"\nSaved top combinations: {args.top_csv}")
    print(f"Saved full report: {args.report_json}")


if __name__ == "__main__":
    main()
